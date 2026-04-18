import sys
from pathlib import Path
import io
import re

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def is_numeric_cell(value: str) -> bool:
    """Return True if the stripped value looks like a number."""
    v = value.strip().replace(',', '').replace('%', '')
    try:
        float(v)
        return True
    except ValueError:
        return False


def cell_real_value(raw):
    """Convert raw cell text to int/float where possible, else return str."""
    if raw is None:
        return ""
    text = str(raw).strip()
    if text == "":
        return ""
    # Try integer first
    try:
        as_float = float(text.replace(',', ''))
        if as_float == int(as_float) and '.' not in text:
            return int(text.replace(',', ''))
        return as_float
    except ValueError:
        return text


# ---------------------------------------------------------------------------
# PDF extraction
# ---------------------------------------------------------------------------

def extract_tables_from_pdf(pdf_path: str):
    """
    Extract tables page by page.

    Returns
    -------
    list of (page_num, list_of_tables)
        where each table is list[list[str|None]]
    """
    tables_by_page = []

    if not HAS_PDFPLUMBER:
        raise SystemExit(
            "pdfplumber is required. Install with: pip install pdfplumber"
        )

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            page_tables = []

            # ---- Strategy 1: pdfplumber structured table extraction ----------
            tables = page.extract_tables({
                "vertical_strategy":   "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance":      3,
                "join_tolerance":      3,
                "edge_min_length":     3,
                "min_words_vertical":  1,
                "min_words_horizontal": 1,
            })

            if tables:
                for tbl in tables:
                    cleaned = _clean_table(tbl)
                    if cleaned:
                        page_tables.append(cleaned)

            # ---- Strategy 2: fallback – explicit rectangles ------------------
            if not page_tables:
                tables = page.extract_tables({
                    "vertical_strategy":   "explicit",
                    "horizontal_strategy": "explicit",
                    "explicit_vertical_lines":   _get_rect_lines(page, 'v'),
                    "explicit_horizontal_lines": _get_rect_lines(page, 'h'),
                })
                if tables:
                    for tbl in tables:
                        cleaned = _clean_table(tbl)
                        if cleaned:
                            page_tables.append(cleaned)

            # ---- Strategy 3: text-based fallback ----------------------------
            if not page_tables:
                text = page.extract_text()
                if text:
                    lines = [l.strip() for l in text.split('\n') if l.strip()]
                    if lines:
                        page_tables.append([line.split() for line in lines])

            if page_tables:
                tables_by_page.append((page_num, page_tables))

    return tables_by_page


def _get_rect_lines(page, direction: str):
    """Collect unique horizontal or vertical line positions from page rects."""
    positions = set()
    for rect in page.rects:
        if direction == 'v':
            positions.add(rect['x0'])
            positions.add(rect['x1'])
        else:
            positions.add(rect['top'])
            positions.add(rect['bottom'])
    return sorted(positions)


def _clean_table(raw_table):
    """Strip whitespace, unify None, remove completely empty rows."""
    cleaned = []
    for row in raw_table:
        new_row = []
        for cell in row:
            if cell is None:
                new_row.append(None)
            else:
                new_row.append(str(cell).strip())
        # Keep row if it has at least one non-empty cell
        if any(c not in (None, "") for c in new_row):
            cleaned.append(new_row)
    return cleaned


# ---------------------------------------------------------------------------
# Merge detection
# ---------------------------------------------------------------------------

def detect_row_merges(row):
    """
    Return list of (start_col_0indexed, span) for cells that should be
    merged horizontally.
    A cell should be merged when it is non-empty and is followed by
    one or more None / empty-string cells.
    """
    merges = []
    i = 0
    while i < len(row):
        val = row[i]
        if val not in (None, ""):
            span = 1
            while i + span < len(row) and row[i + span] in (None, ""):
                span += 1
            if span > 1:
                merges.append((i, span))
            i += span
        else:
            i += 1
    return merges


def detect_col_merges(table):
    """
    Return dict  {(row_idx_0based, col_idx_0based): row_span}
    for cells that should be merged vertically (same value, consecutive rows,
    only when the column has None beneath it).
    """
    if not table:
        return {}

    max_col = max(len(r) for r in table)
    col_merges = {}

    for c in range(max_col):
        r = 0
        while r < len(table):
            val = table[r][c] if c < len(table[r]) else None
            if val not in (None, ""):
                span = 1
                while (r + span < len(table) and
                       (c >= len(table[r + span]) or
                        table[r + span][c] in (None, ""))):
                    span += 1
                if span > 1:
                    col_merges[(r, c)] = span
                r += span
            else:
                r += 1
    return col_merges


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def apply_table_to_worksheet(ws, start_row: int, table):
    """
    Write *table* into *ws* beginning at *start_row*.
    Handles:
      - thin borders on every cell
      - right-align for numeric cells, center for headers, left for text
      - horizontal merging (None/empty after a value)
      - vertical merging (None/empty below a value in same column)
      - bold header row when it is the result of a full-width merge
    """
    if not table:
        return start_row

    thin = make_thin_border()
    max_col = max(len(r) for r in table)
    num_rows = len(table)

    # --- Detect merges -------------------------------------------------------
    col_merge_map = detect_col_merges(table)   # {(r,c): rowspan}
    used_by_vmerge = set()                      # cells already covered

    # Pre-build set of cells covered by vertical merges
    for (r, c), span in col_merge_map.items():
        for dr in range(1, span):
            used_by_vmerge.add((r + dr, c))

    applied_merges = set()   # track ws.merge_cells calls to avoid duplicates

    # --- Write data + style --------------------------------------------------
    for r_idx, row in enumerate(table):
        excel_row = start_row + r_idx

        # Detect horizontal merges for this row
        h_merges = detect_row_merges(row)
        h_merge_map = {}          # start_col_0idx -> span
        h_covered = set()
        for (sc, span) in h_merges:
            h_merge_map[sc] = span
            for dc in range(1, span):
                h_covered.add(sc + dc)

        # Is this a header row? (single value spanning all columns)
        is_header_row = (
            len(h_merges) == 1 and
            h_merges[0][0] == 0 and
            h_merges[0][1] == max_col
        )

        c_idx = 0
        while c_idx < max_col:
            excel_col = c_idx + 1
            raw = row[c_idx] if c_idx < len(row) else None
            value = cell_real_value(raw)

            cell = ws.cell(row=excel_row, column=excel_col)
            cell.border = thin

            # Skip cells that are covered by a horizontal merge
            if c_idx in h_covered:
                c_idx += 1
                continue

            cell.value = value

            # Alignment
            if is_header_row:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
                cell.font = Font(bold=True)
            elif isinstance(value, (int, float)):
                cell.alignment = Alignment(
                    horizontal='right', vertical='center'
                )
            elif value == "":
                cell.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
            else:
                # Heuristic: short all-caps or title-like → center
                s = str(value)
                if len(s) < 30 and (s.isupper() or s.istitle()):
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center'
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=True
                    )

            # --- Apply horizontal merge --------------------------------------
            h_span = h_merge_map.get(c_idx, 1)
            v_span = col_merge_map.get((r_idx, c_idx), 1)

            if h_span > 1 or v_span > 1:
                merge_key = (excel_row, excel_col,
                             excel_row + v_span - 1,
                             excel_col + h_span - 1)
                if merge_key not in applied_merges:
                    ws.merge_cells(
                        start_row=excel_row,
                        start_column=excel_col,
                        end_row=excel_row + v_span - 1,
                        end_column=excel_col + h_span - 1,
                    )
                    applied_merges.add(merge_key)
                    # Fill borders on all covered cells of this merge block
                    for dr in range(v_span):
                        for dc in range(h_span):
                            ws.cell(
                                row=excel_row + dr,
                                column=excel_col + dc
                            ).border = thin

            c_idx += h_span

    return start_row + num_rows


def auto_adjust_column_widths(ws, min_width=6, max_width=60):
    """Set column widths based on maximum content length."""
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            length = len(str(cell.value))
            # Account for merged cells (don't over-expand)
            if hasattr(cell, 'column') and length > 0:
                col_widths[col] = max(col_widths.get(col, min_width), length + 2)

    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = min(width, max_width)


def write_tables_to_worksheet(ws, tables, start_row: int) -> int:
    """Write all *tables* into *ws*, separated by 2 blank rows."""
    current_row = start_row
    for table in tables:
        current_row = apply_table_to_worksheet(ws, current_row, table)
        current_row += 2   # blank-row gap between tables
    return current_row


# ---------------------------------------------------------------------------
# Workbook creation
# ---------------------------------------------------------------------------

def create_excel_single_sheet(tables_by_page: list, output_path: str):
    if not HAS_OPENPYXL:
        raise SystemExit(
            "openpyxl is required. Install with: pip install openpyxl"
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "All Tables"

    current_row = 1
    for _page_num, tables in tables_by_page:
        current_row = write_tables_to_worksheet(ws, tables, current_row)

    auto_adjust_column_widths(ws)
    wb.save(output_path)
    print(f"Saved (single sheet): {output_path}")


def create_excel_multiple_sheets(tables_by_page: list, output_path: str):
    if not HAS_OPENPYXL:
        raise SystemExit(
            "openpyxl is required. Install with: pip install openpyxl"
        )
    wb = Workbook()
    wb.remove(wb.active)

    for page_num, tables in tables_by_page:
        ws = wb.create_sheet(title=f"Page {page_num}")
        write_tables_to_worksheet(ws, tables, 1)
        auto_adjust_column_widths(ws)

    wb.save(output_path)
    print(f"Saved (multiple sheets): {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    if len(sys.argv) < 4:
        raise SystemExit(
            "Usage: pdf_to_xlsx.py <input-pdf> <output-xlsx> <layout>\n"
            "  layout: single | multiple"
        )

    input_path  = Path(sys.argv[1]).resolve()
    output_path = Path(sys.argv[2]).resolve()
    layout      = sys.argv[3].lower()

    if layout not in ("single", "multiple"):
        raise SystemExit("Layout must be 'single' or 'multiple'.")

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        tables_by_page = extract_tables_from_pdf(str(input_path))

        if not tables_by_page:
            print("Warning: no tables detected in PDF.", file=sys.stderr)
            return 1

        if layout == "single":
            create_excel_single_sheet(tables_by_page, str(output_path))
        else:
            create_excel_multiple_sheets(tables_by_page, str(output_path))

        return 0

    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())